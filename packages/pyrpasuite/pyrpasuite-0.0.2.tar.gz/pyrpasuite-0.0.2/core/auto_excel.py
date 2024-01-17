import openpyxl

class AutoExcel():
    """
    A class to automate Excel operations using openpyxl.
    """

    def __init__(self, path):
        """
        Initialize AutoExcel with the path of the Excel file.

        :param path: The path of the Excel file.
        """
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb.active

    def read_cell(self, cell):
        """
        Read the value of a specific cell.

        :param cell: The cell to read.
        :return: The value of the cell.
        """
        return self.ws[cell].value

    def write_cell(self, cell, value):
        """
        Write a value to a specific cell.

        :param cell: The cell to write to.
        :param value: The value to write.
        """
        self.ws[cell].value = value

    def read_row(self, row):
        """
        Read all values in a specific row.

        :param row: The row to read.
        :return: A list of values in the row.
        """
        return [cell.value for cell in self.ws[row]]

    def write_row(self, row, values):
        """
        Write a list of values to a specific row.

        :param row: The row to write to.
        :param values: The list of values to write.
        """
        for cell, value in zip(self.ws[row], values):
            cell.value = value

    def add_sheet(self, title):
        """
        Add a new sheet with the given title.

        :param title: The title of the new sheet.
        """
        self.wb.create_sheet(title)

    def remove_sheet(self, title):
        """
        Remove a sheet with the given title.

        :param title: The title of the sheet to remove.
        """
        self.wb.remove(self.wb[title])

    def switch_sheet(self, title):
        """
        Switch to another sheet with the given title.

        :param title: The title of the sheet to switch to.
        """
        self.ws = self.wb[title]

    def find_cells(self, value):
        """
        Find all cells with the given value.

        :param value: The value to find.
        :return: A list of cells with the given value.
        """
        return [cell for row in self.ws.iter_rows() for cell in row if cell.value == value]

    def save_and_close(self, filename="new_file.xlsm"):
        """
        Save and close the workbook. If no filename is given, it defaults to 'new_file.xlsm'.

        :param filename: The filename to save as. Defaults to "new_file.xlsm".
        """
        self.wb.save(filename)
        self.wb.close()

import openpyxl

TESTRESULT = 3
EXPECTED = 2
SIMPLE = 1
def get_example_list(excelFile, sheetName = None, headerMatcher = None, headerUnmatcher = None):
    wb = openpyxl.load_workbook(excelFile)    
    # Define variable to read sheet
    if sheetName == None:
        ws = wb[wb.sheetnames[0]]
    else:
        if sheetName in wb.sheetnames:
            ws = wb[sheetName]
        else:
            raise Exception(sheetName + " sheet does not exist.")        
    
    max_row = ws.max_row
    max_column = ws.max_column
    # Iterate the loop to read the cell values
    IsFound = False
    parameterRow = 0
    parameterCol = 0
    for row in range(1, max_row):
        for col in range(2, max_column):
            if "Parameter Name" in str(ws.cell(row, col).value) :
                parameterRow = row
                parameterCol = col
                IsFound = True
                break
        if IsFound == True:
            break
    if IsFound == False:
        raise Exception("Paramter Name grid is not found.")
    # print(parameterRow)
    # print(parameterCol)

    if str(ws.cell(parameterRow, parameterCol+3).value) == "Test Result":
        columnStep = TESTRESULT
        actualParameterRow = parameterRow -1
    elif str(ws.cell(parameterRow, parameterCol+2).value) == "Expected":
        columnStep = EXPECTED
        actualParameterRow = parameterRow -1
    else:
        columnStep = SIMPLE
        actualParameterRow = parameterRow

    parameterNames = "HeaderName"
    for row in range(parameterRow + 1, max_row +1):
        if ws.cell(row, parameterCol).value != None and ws.cell(row, parameterCol).value != "NA" :
            parameterName =  str((ws.cell(row, parameterCol).value))
            parameterNames = parameterNames + ", " + parameterName
            if columnStep > SIMPLE:
                parameterNames = parameterNames + ", " + parameterName + "Expected"
                if columnStep == TESTRESULT:
                    parameterNames = parameterNames + ", " + parameterName + "TestResult"

    print("The parameter names are " + parameterNames)
    
    testcaseSetList = []
    # print('max_column', max_column)
    for col in range(parameterCol+1, max_column +1, columnStep):
        header = str(ws.cell(actualParameterRow, col).value)
        if header == None or header == 'None':
            break
        if headerMatcher != None and headerMatcher not in header :
            continue
        if headerUnmatcher != None and headerUnmatcher in header :
            continue
        testcaseSet = []
        testcaseSet.append(ws.cell(actualParameterRow, col).value)
        for row in range(parameterRow + 1, max_row +1):
            if ws.cell(row, parameterCol).value != None and ws.cell(row, parameterCol).value != "NA" :
                testcaseSet.append(ws.cell(row, col).value)
                if columnStep > SIMPLE:
                    testcaseSet.append(ws.cell(row, col+1).value)
                    if columnStep == TESTRESULT:
                        testcaseSet.append(ws.cell(row, col+2).value)
        testcaseSetList.append(testcaseSet)

    return testcaseSetList

def get_example_table(excelFile,sheetName = None,headerRow = 1,startColumn = '`'):
    nStartCol = 1
    if startColumn != '`':
        try:    
            nStartCol = ord(startColumn) - 64
            searchStartCol = nStartCol
            IsBeforeStartColumn = False
        except:
            raise Exception("Start Column must in A~Z")
        if nStartCol < 1 or nStartCol > 26:
            raise Exception("Start Column must in A~Z")
    else:
        searchStartCol = 1
        IsBeforeStartColumn = True

    wb = openpyxl.load_workbook(excelFile)    
    # Define variable to read sheet
    if sheetName == None:
        ws = wb[wb.sheetnames[0]]
    else:
        if sheetName in wb.sheetnames:
            ws = wb[sheetName]
        else:
            raise Exception(sheetName + " sheet does not exist.")

    maxTableCol =  0
    parameterNames = "Table Parameter Names are below"
    for col in range(searchStartCol, ws.max_column + 1):
        headerValue = ws.cell(headerRow, col).value
        if headerValue is None and IsBeforeStartColumn:
            continue
        elif headerValue is not None and IsBeforeStartColumn:
            nStartCol = col
            IsBeforeStartColumn = False
            parameterNames = parameterNames + ", " + str(headerValue)
        elif headerValue is None and not IsBeforeStartColumn:
            maxTableCol = col - 1
            break
        else:
            parameterNames = parameterNames + ", " + str(headerValue)

    if maxTableCol == 0:
        maxTableCol =  ws.max_column

    print(parameterNames)

    testcaseSetList = []
    for row in range( headerRow + 1, ws.max_row + 1):
        if ws.cell(row, nStartCol).value == None:
            break
        testcaseSet = []
        for col in range(nStartCol, maxTableCol + 1):
            testcaseSet.append(ws.cell(row, col).value)
        testcaseSetList.append(testcaseSet)

    return testcaseSetList
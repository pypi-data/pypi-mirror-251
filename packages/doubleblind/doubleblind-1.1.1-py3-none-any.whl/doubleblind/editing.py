from pathlib import Path

from openpyxl import load_workbook


def get_mod_filename(filename: Path):
    return filename.parent.joinpath(f"{filename.stem}_unblinded{filename.suffix}")


def edit_excel(filename: Path, decode_dict: dict):
    was_modified = False
    # Load the workbook
    wb = load_workbook(filename)

    # Iterate over all sheets
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        # Find and replace strings
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    old_value = cell.value
                    new_value = old_value
                    for coded, raw in decode_dict.items():
                        new_value = new_value.replace(coded, raw)
                    if new_value != old_value:
                        was_modified = True
                        cell.value = new_value

    # Save the modified workbook
    if was_modified:
        mod_filename = get_mod_filename(filename)
        wb.save(mod_filename)
        return mod_filename


def edit_text(filename: Path, decode_dict: dict):
    with open(filename) as infile:
        text = infile.read()
        mod_text = text
    for coded, raw in decode_dict.items():
        mod_text = mod_text.replace(coded, raw)

    was_modified = text != mod_text

    if was_modified:
        mod_filename = get_mod_filename(filename)
        with open(mod_filename, 'w') as outfile:
            outfile.write(mod_text)
        return mod_filename

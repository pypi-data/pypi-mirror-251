from doubleblind.editing import *
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from doubleblind.editing import *


@pytest.fixture
def sample_excel_file(tmp_path):
    # Create a sample Excel file for testing
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "Hello"
    ws['A2'] = "World"
    ws['B1'] = "Test"
    ws['B2'] = "Data"
    ws['A1'].font = Font(bold=True)
    ws['A2'].font = Font(bold=True)
    wb.save(tmp_path / "sample.xlsx")
    return tmp_path / "sample.xlsx"


@pytest.fixture
def sample_text_file(tmp_path):
    # Create a sample text file for testing
    text = "Hello, World! This is a test file."
    file_path = tmp_path / "sample.txt"
    with open(file_path, 'w') as f:
        f.write(text)
    return file_path


@pytest.mark.parametrize('in_path,truth', [
    ("path/to/file.xlsx", "path/to/file_unblinded.xlsx"),
    ('file_1.txt', 'file_1_unblinded.txt'),
    ('path to/dir/some otherfile123.csv', 'path to/dir/some otherfile123_unblinded.csv')
])
def test_get_mod_filename(in_path, truth):
    # Test the get_mod_filename function
    file_path = Path(in_path)
    mod_file_path = get_mod_filename(file_path)
    assert mod_file_path == Path(truth)


def test_edit_excel_with_modifications(sample_excel_file):
    # Test the edit_excel function with modifications in the Excel file
    decode_dict = {
        "Hello": "Hi",
        "Data": "Information"
    }
    mod_file_path = edit_excel(sample_excel_file, decode_dict)
    assert mod_file_path is not None
    assert mod_file_path.exists()
    wb = load_workbook(mod_file_path)
    ws = wb.active
    assert ws['A1'].value == "Hi"
    assert ws['B2'].value == "Information"


def test_edit_excel_without_modifications(sample_excel_file):
    # Test the edit_excel function without any modifications in the Excel file
    decode_dict = {
        "test1": "example",
        "hello": "Hi"
    }
    mod_file_path = edit_excel(sample_excel_file, decode_dict)
    assert mod_file_path is None


def test_edit_text_with_modifications(sample_text_file):
    # Test the edit_text function with modifications in the text file
    decode_dict = {
        "Hello": "Hi",
        "test": "example"
    }
    mod_file_path = edit_text(sample_text_file, decode_dict)
    assert mod_file_path is not None
    assert mod_file_path.exists()
    with open(mod_file_path, 'r') as f:
        mod_text = f.read()
    assert mod_text == "Hi, World! This is a example file."


def test_edit_text_without_modifications(sample_text_file):
    # Test the edit_text function without any modifications in the text file
    decode_dict = {
        "test1": "example",
        "hello": "Hi"
    }
    mod_file_path = edit_text(sample_text_file, decode_dict)
    assert mod_file_path is None

from openpyxl import Workbook, load_workbook


class ExcelUtil:

    @classmethod
    def create_excel(cls, file_name: str, titles: list, data: list, sheet_name: str = 'Sheet1'):
        if not file_name.endswith('.xlsx'):
            raise ValueError('file_name must endswith .xlsx')
        if len(titles) != len(data):
            raise ValueError('titles and data length must be same')
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        new_data = [titles, ] + data
        for x, row_value in enumerate(new_data, start=1):
            for y, value in enumerate(row_value, start=1):
                cell = ws.cell(row=x, column=y)
                cell.value = value
        wb.save(file_name)
        return file_name

    @classmethod
    def open_excel(cls, file_name, data_only: bool = True):
        wb = load_workbook(file_name, data_only=data_only)
        return wb


if __name__ == '__main__':
    titles = ['company', 'email']
    data = [['com1', 'email1'], ['com2', 'email2']]
    file = ExcelUtil.create_excel('aaa.xlsx', titles, data)
